from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Any

from openpyxl import load_workbook
from rapidfuzz import fuzz
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.models import (
    Category,
    ChannelShareLink,
    ImportBatch,
    ImportError,
    ImportRawRow,
    Provider,
    Resource,
    ResourceChannel,
)
from app.providers import registry, url_hash
from app.services.resources import create_resource
from app.services.category_governance import resolve_categories
from app.services.text import clean_isbn, normalize_title


HEADER_ALIASES = {
    "书名": "title",
    "名称": "title",
    "title": "title",
    "作者": "author",
    "author": "author",
    "副标题": "subtitle",
    "subtitle": "subtitle",
    "isbn": "isbn",
    "ISBN": "isbn",
    "出版方": "publisher",
    "出版社": "publisher",
    "publisher": "publisher",
    "出版年份": "publish_year",
    "出版年": "publish_year",
    "年份": "publish_year",
    "publish_year": "publish_year",
    "简介": "description",
    "内容简介": "description",
    "description": "description",
    "语言": "language",
    "language": "language",
    "分类": "category",
    "category": "category",
    "网盘链接": "share_url",
    "链接": "share_url",
    "share_url": "share_url",
    "提取码": "extract_code",
    "密码": "extract_code",
    "extract_code": "extract_code",
    "格式": "formats",
    "formats": "formats",
    "版权状态": "copyright_status",
    "copyright_status": "copyright_status",
    "授权说明": "source_reference",
    "source_reference": "source_reference",
}


@dataclass(slots=True)
class CommitResult:
    committed: int
    skipped: int


def parse_upload(filename: str, content: bytes) -> list[dict[str, Any]]:
    if len(content) > get_settings().import_max_bytes:
        raise ValueError("文件超过 5MB 限制")
    suffix = filename.rsplit(".", 1)[-1].casefold() if "." in filename else ""
    if suffix == "csv":
        return _parse_csv(content)
    if suffix == "xlsx":
        return _parse_xlsx(content)
    raise ValueError("仅支持 .xlsx 或 .csv 文件")


def create_preview(db: Session, filename: str, content: bytes, admin_id: int) -> ImportBatch:
    rows = parse_upload(filename, content)
    limit = get_settings().import_max_rows
    if len(rows) > limit:
        raise ValueError(f"一次最多导入 {limit} 行")
    batch = ImportBatch(original_filename=filename[:255], created_by_id=admin_id, total_rows=len(rows))
    db.add(batch)
    db.flush()
    seen_hashes: dict[str, int] = {}
    existing_links = {
        item.normalized_url_hash: item
        for item in db.scalars(select(ChannelShareLink)).all()
    }
    resources = list(db.scalars(select(Resource)))
    for number, raw in enumerate(rows, start=2):
        parsed, status, message, digest, matched_id = _evaluate_row(
            raw, seen_hashes, existing_links, resources, number
        )
        db.add(
            ImportRawRow(
                batch_id=batch.id,
                row_number=number,
                raw_data=raw,
                parsed_data=parsed,
                row_status=status,
                message=message,
                normalized_url_hash=digest,
                matched_resource_id=matched_id,
            )
        )
        if status == "ready":
            batch.ready_rows += 1
        elif status == "warning":
            batch.warning_rows += 1
        elif status.startswith("duplicate"):
            batch.duplicate_rows += 1
        elif status == "conflict":
            batch.conflict_rows += 1
        else:
            batch.error_rows += 1
            db.add(
                ImportError(
                    batch_id=batch.id,
                    row_number=number,
                    error_code="invalid_row",
                    message=message or "数据不完整",
                )
            )
    db.commit()
    db.refresh(batch)
    return batch


def commit_preview(db: Session, batch: ImportBatch, selected_row_ids: set[int]) -> CommitResult:
    if batch.status != "preview":
        raise ValueError("该批次已经处理，不能重复提交")
    committed = 0
    skipped = 0
    for row in batch.rows:
        if row.id not in selected_row_ids or row.row_status not in {"ready", "warning"}:
            skipped += 1
            continue
        if db.scalar(
            select(ChannelShareLink.id).where(
                ChannelShareLink.normalized_url_hash == row.normalized_url_hash
            )
        ):
            row.row_status = "duplicate_existing"
            row.message = "提交时发现链接已存在，已跳过"
            skipped += 1
            continue
        data = row.parsed_data
        resource = db.get(Resource, row.matched_resource_id) if row.matched_resource_id else None
        if not resource:
            resource = create_resource(
                db,
                {
                    "title": data["title"],
                    "subtitle": data.get("subtitle"),
                    "author": data.get("author"),
                    "isbn": data.get("isbn"),
                    "publisher": data.get("publisher"),
                    "publish_year": data.get("publish_year"),
                    "description": data.get("description"),
                    "language": data.get("language") or "zh-CN",
                    "formats": data.get("formats"),
                    "copyright_status": data.get("copyright_status") or "pending",
                    "source_reference": data.get("source_reference"),
                    "publish_status": "draft",
                },
            )
            category_name = data.get("category")
            if category_name:
                resource.source_category_main = category_name
                try:
                    resource.categories = resolve_categories(db, category_name)
                except ValueError as exc:
                    row.message = str(exc)
        provider = db.scalar(select(Provider).where(Provider.code == data["provider_code"]))
        channel = db.scalar(
            select(ResourceChannel).where(
                ResourceChannel.resource_id == resource.id,
                ResourceChannel.provider_id == provider.id,
            )
        )
        if not channel:
            channel = ResourceChannel(resource_id=resource.id, provider_id=provider.id, status="active")
            db.add(channel)
            db.flush()
        db.add(
            ChannelShareLink(
                channel_id=channel.id,
                provider_id=provider.id,
                provider_share_id=data["provider_share_id"],
                share_url=data["normalized_url"],
                normalized_url=data["normalized_url"],
                normalized_url_hash=row.normalized_url_hash,
                extract_code=data.get("extract_code"),
                status="pending",
                is_visible=False,
            )
        )
        row.row_status = "committed"
        committed += 1
    batch.status = "committed"
    batch.committed_rows = committed
    from app.models.base import utcnow

    batch.committed_at = utcnow()
    db.commit()
    return CommitResult(committed=committed, skipped=skipped)


def _evaluate_row(
    raw: dict[str, Any],
    seen_hashes: dict[str, int],
    existing_links: dict[str, ChannelShareLink],
    resources: list[Resource],
    row_number: int,
) -> tuple[dict[str, Any], str, str | None, str | None, int | None]:
    data = {_canonical_header(key): _cell(value) for key, value in raw.items() if _canonical_header(key)}
    title = data.get("title", "").strip()
    url = data.get("share_url", "").strip()
    if not title or not url:
        return data, "error", "书名和网盘链接为必填项", None, None
    try:
        parsed_link = registry.recognize(url, data.get("extract_code"))
    except ValueError as exc:
        return data, "error", str(exc), None, None
    digest = url_hash(parsed_link.normalized_url)
    data.update(
        {
            "title": title,
            "normalized_title": normalize_title(title),
            "isbn": clean_isbn(data.get("isbn")),
            "provider_code": parsed_link.provider_code,
            "provider_share_id": parsed_link.share_id,
            "normalized_url": parsed_link.normalized_url,
            "extract_code": parsed_link.extract_code,
        }
    )
    if digest in seen_hashes:
        return data, "duplicate_batch", f"与本批次第 {seen_hashes[digest]} 行链接重复", digest, None
    seen_hashes[digest] = row_number
    title_match = next((item for item in resources if item.normalized_title == data["normalized_title"]), None)
    existing_link = existing_links.get(digest)
    if existing_link:
        channel = existing_link.channel
        if channel and channel.resource.normalized_title == data["normalized_title"]:
            return data, "duplicate_existing", "该书已存在相同链接", digest, channel.resource_id
        linked_title = channel.resource.title if channel and channel.resource else "其他资源"
        return data, "conflict", f"危险：该链接已属于《{linked_title}》，禁止导入", digest, None
    if title_match:
        return data, "ready", f"将补充到现有资源《{title_match.title}》", digest, title_match.id
    if resources:
        candidate = max(resources, key=lambda item: fuzz.ratio(data["normalized_title"], item.normalized_title))
        score = fuzz.ratio(data["normalized_title"], candidate.normalized_title)
        if score >= 86:
            return data, "warning", f"可能与《{candidate.title}》重复（相似度 {score}%），请人工确认", digest, None
    return data, "ready", "将新建资源，链接待检测后才会前台显示", digest, None


def _parse_csv(content: bytes) -> list[dict[str, Any]]:
    text = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("CSV 编码无法识别，请保存为 UTF-8")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValueError("表格缺少表头")
    return [dict(row) for row in reader if any(_cell(value) for value in row.values())]


def _parse_xlsx(content: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError("Excel 文件无法读取") from exc
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = [_cell(value) for value in next(iterator)]
    except StopIteration as exc:
        raise ValueError("Excel 文件为空") from exc
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        for row in iterator
        if any(_cell(value) for value in row)
    ]


def _canonical_header(value: str) -> str | None:
    key = str(value or "").strip()
    return HEADER_ALIASES.get(key) or HEADER_ALIASES.get(key.casefold())


def _cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()
